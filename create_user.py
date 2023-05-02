# Import required Libraries
# Snowpark
from snowflake.snowpark.session import Session
# Pandas
import pandas as pd
# streamlit
import streamlit as st
# excel
from openpyxl import reader, load_workbook, Workbook
# etc
import io

def main():
    #st.title = 'Create Snowflake User'
    #st.snow()

    # ユーザーリストアップロード
    df = upload()

    if df is not None:
        # ログイン情報入力
        connection_parameters = input_login_info()
        if connection_parameters:
            if st.button('Create user', on_click=None):
                session = Session.builder.configs(connection_parameters).create()
                for row in df.iterrows():
                    user = row[1][0]
                    st.write(user)
                    a = session.sql(f"CREATE USER {user} PASSWORD='abc123' DEFAULT_ROLE = sysadmin DEFAULT_SECONDARY_ROLES = ('ALL') MUST_CHANGE_PASSWORD = TRUE").collect()
                st.write('Created users.')
                session.close()

def upload():
    #アップローダー表示
    st.title('Create a new user in Snowflake')
    upload_file=st.file_uploader('upload user list file', type='xlsx')

    if upload_file:
        wb=Workbook()

        #アップロードファイルを読み込み
        wb=load_workbook(upload_file,read_only=False)
        buffer=io.BytesIO()
        wb.save(buffer)

        #アップロードファイル内容表示
        df=pd.read_excel(upload_file,sheet_name=wb.sheetnames[0])
        st.write(df)

        return df


#ダウンロード
def download(buffer):
    st.download_button(
        label='Download Excel worksheet without index'
        ,data=buffer
        ,file_name='fk.xlsx'
    )

# ユーザー作成
def input_login_info():
    account = st.text_input('account', '')
    user = st.text_input('user', '')
    password = st.text_input('password', '')
    role = st.text_input('role', '')
    warehouse = st.text_input('warehouse', '')
    database = st.text_input('database', '')
    schema = st.text_input('schema', '')

    if account and user and password and role and warehouse and database and schema:
        connection_parameters = {
        "account": account,
        "user": user,
        "password": password,
        "role": role,
        "warehouse": warehouse,
        "database": database,
        "schema": schema 
        }
        return connection_parameters

main()
